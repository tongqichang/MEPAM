import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def convert_to_numeric_ids_by_column(input_file_path, sheet_name):
    # 读取Excel文件
    wb = load_workbook(input_file_path)

    if sheet_name not in wb.sheetnames:
        print(f"错误：工作表 '{sheet_name}' 不存在!")
        return

    ws = wb[sheet_name]

    # 创建新工作表（在原名称后加 '_id'）
    new_sheet_name = f"{sheet_name}_id"
    if new_sheet_name in wb.sheetnames:
        wb.remove(wb[new_sheet_name])
    new_ws = wb.create_sheet(new_sheet_name)

    # 复制标题行
    for col in range(1, ws.max_column + 1):
        new_ws.cell(row=1, column=col, value=ws.cell(row=1, column=col).value)

    # 值到ID的映射字典
    value_to_id = {}
    current_id = 1

    # 按列处理数据（跳过标题行）
    for col in range(1, ws.max_column + 1):
        for row in range(2, ws.max_row + 1):
            original_value = ws.cell(row=row, column=col).value
            new_value = None  # 默认处理空单元格

            if original_value is not None:
                if original_value not in value_to_id:
                    value_to_id[original_value] = current_id
                    current_id += 1
                new_value = value_to_id[original_value]

            # 写入新工作表
            new_ws.cell(row=row, column=col, value=new_value)

    # 保存修改（不覆盖原sheet）
    wb.save(input_file_path)
    print(f"处理完成！已创建新工作表 '{new_sheet_name}'（按列顺序分配ID）")


if __name__ == "__main__":
    input_file = r"C:\Users\LEGION\Desktop\第六部分网络图\纤维素酶produce_培养基.xlsx"
    target_sheet = "cellulase_produce"

    convert_to_numeric_ids_by_column(input_file, target_sheet)